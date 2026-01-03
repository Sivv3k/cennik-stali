"""Serwis do eksportu danych cennikowych do Excel/CSV."""

from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from ..models import (
    Material,
    MaterialCategory,
    BasePrice,
    GrindingPrice,
    GrindingProvider,
    FilmPrice,
    FilmType,
    ThicknessModifier,
    WidthModifier,
)


class PriceExporter:
    """Eksporter danych cennikowych do Excel/CSV."""

    # Style dla Excel
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    def __init__(self, db: Session):
        self.db = db

    def export_base_prices(
        self,
        categories: Optional[list[str]] = None,
        thickness_min: Optional[float] = None,
        thickness_max: Optional[float] = None,
        width_min: Optional[float] = None,
        width_max: Optional[float] = None,
        surface_finishes: Optional[list[str]] = None,
        only_active: bool = True,
    ) -> bytes:
        """Eksportuj ceny bazowe do Excel.

        Args:
            categories: Lista kategorii materialow (stal_nierdzewna, stal_czarna, aluminium)
            thickness_min: Minimalna grubosc
            thickness_max: Maksymalna grubosc
            width_min: Minimalna szerokosc
            width_max: Maksymalna szerokosc
            surface_finishes: Lista wykonczn (2B, BA, 1D, itp.)
            only_active: Tylko aktywne ceny

        Returns:
            bytes: Zawartosc pliku Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Ceny bazowe"

        # Naglowki
        headers = [
            "Gatunek", "Nazwa materialu", "Kategoria", "Wykoncznie",
            "Grubosc (mm)", "Szerokosc (mm)", "Dlugosc (mm)", "Cena PLN/kg", "Uwagi"
        ]
        self._write_headers(ws, headers)

        # Pobierz dane
        query = self.db.query(BasePrice).options(joinedload(BasePrice.material))

        if only_active:
            query = query.filter(BasePrice.is_active == True)

        if categories:
            category_enums = [MaterialCategory(c) for c in categories]
            query = query.join(Material).filter(Material.category.in_(category_enums))

        if thickness_min is not None:
            query = query.filter(BasePrice.thickness >= thickness_min)
        if thickness_max is not None:
            query = query.filter(BasePrice.thickness <= thickness_max)

        if width_min is not None:
            query = query.filter(BasePrice.width >= width_min)
        if width_max is not None:
            query = query.filter(BasePrice.width <= width_max)

        if surface_finishes:
            query = query.filter(BasePrice.surface_finish.in_(surface_finishes))

        prices = query.order_by(
            BasePrice.material_id,
            BasePrice.surface_finish,
            BasePrice.thickness,
            BasePrice.width
        ).all()

        # Zapisz dane
        for row_idx, price in enumerate(prices, start=2):
            material = price.material
            ws.cell(row=row_idx, column=1, value=material.grade)
            ws.cell(row=row_idx, column=2, value=material.name)
            ws.cell(row=row_idx, column=3, value=material.category.value)
            ws.cell(row=row_idx, column=4, value=price.surface_finish)
            ws.cell(row=row_idx, column=5, value=price.thickness)
            ws.cell(row=row_idx, column=6, value=price.width)
            ws.cell(row=row_idx, column=7, value=price.length)
            ws.cell(row=row_idx, column=8, value=price.price_pln_per_kg)
            ws.cell(row=row_idx, column=9, value=price.notes or "")

            # Formatowanie
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.BORDER
                if col in [5, 6, 7, 8]:  # Kolumny numeryczne
                    cell.alignment = self.CENTER_ALIGN

        # Autofit kolumn
        self._autofit_columns(ws)

        # Zapisz do BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_grinding_prices(
        self,
        providers: Optional[list[str]] = None,
        thickness_min: Optional[float] = None,
        thickness_max: Optional[float] = None,
        only_active: bool = True,
    ) -> bytes:
        """Eksportuj ceny szlifowania do Excel.

        Args:
            providers: Lista dostawcow (CAMU, BABCIA, COSTA, BORYS)
            thickness_min: Minimalna grubosc
            thickness_max: Maksymalna grubosc
            only_active: Tylko aktywne ceny

        Returns:
            bytes: Zawartosc pliku Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Cennik szlifu"

        # Naglowki
        headers = [
            "Dostawca", "Granulacja", "Grubosc (mm)", "Cena PLN/kg", "Z SB", "Wariant szerokosci"
        ]
        self._write_headers(ws, headers)

        # Pobierz dane
        query = self.db.query(GrindingPrice)

        if only_active:
            query = query.filter(GrindingPrice.is_active == True)

        if providers:
            provider_enums = [GrindingProvider(p) for p in providers]
            query = query.filter(GrindingPrice.provider.in_(provider_enums))

        if thickness_min is not None:
            query = query.filter(GrindingPrice.thickness >= thickness_min)
        if thickness_max is not None:
            query = query.filter(GrindingPrice.thickness <= thickness_max)

        prices = query.order_by(
            GrindingPrice.provider,
            GrindingPrice.grit,
            GrindingPrice.thickness
        ).all()

        # Zapisz dane
        for row_idx, price in enumerate(prices, start=2):
            ws.cell(row=row_idx, column=1, value=price.provider.value)
            ws.cell(row=row_idx, column=2, value=price.grit)
            ws.cell(row=row_idx, column=3, value=price.thickness)
            ws.cell(row=row_idx, column=4, value=price.price_pln_per_kg)
            ws.cell(row=row_idx, column=5, value="Tak" if price.with_sb else "Nie")
            ws.cell(row=row_idx, column=6, value=price.width_variant or "")

            # Formatowanie
            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.BORDER
                if col in [3, 4]:
                    cell.alignment = self.CENTER_ALIGN

        self._autofit_columns(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_film_prices(
        self,
        film_types: Optional[list[str]] = None,
        thickness_min: Optional[float] = None,
        thickness_max: Optional[float] = None,
        only_active: bool = True,
    ) -> bytes:
        """Eksportuj ceny folii do Excel.

        Args:
            film_types: Lista typow folii
            thickness_min: Minimalna grubosc
            thickness_max: Maksymalna grubosc
            only_active: Tylko aktywne ceny

        Returns:
            bytes: Zawartosc pliku Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Cennik folii"

        # Naglowki
        headers = ["Typ folii", "Grubosc (mm)", "Cena PLN/kg"]
        self._write_headers(ws, headers)

        # Pobierz dane
        query = self.db.query(FilmPrice)

        if only_active:
            query = query.filter(FilmPrice.is_active == True)

        if film_types:
            type_enums = [FilmType(t) for t in film_types]
            query = query.filter(FilmPrice.film_type.in_(type_enums))

        if thickness_min is not None:
            query = query.filter(FilmPrice.thickness >= thickness_min)
        if thickness_max is not None:
            query = query.filter(FilmPrice.thickness <= thickness_max)

        prices = query.order_by(
            FilmPrice.film_type,
            FilmPrice.thickness
        ).all()

        # Zapisz dane
        for row_idx, price in enumerate(prices, start=2):
            ws.cell(row=row_idx, column=1, value=price.film_type.value)
            ws.cell(row=row_idx, column=2, value=price.thickness)
            ws.cell(row=row_idx, column=3, value=price.price_pln_per_kg)

            for col in range(1, 4):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.BORDER
                if col in [2, 3]:
                    cell.alignment = self.CENTER_ALIGN

        self._autofit_columns(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_modifiers(self) -> bytes:
        """Eksportuj modyfikatory cen do Excel.

        Returns:
            bytes: Zawartosc pliku Excel
        """
        wb = Workbook()

        # Arkusz 1: Modyfikatory grubosci
        ws1 = wb.active
        ws1.title = "Modyfikatory grubosci"

        headers = ["Gatunek", "Wykoncznie", "Szerokosc bazowa", "Grubosc (mm)", "Modyfikator PLN/kg"]
        self._write_headers(ws1, headers)

        thickness_mods = self.db.query(ThicknessModifier).order_by(
            ThicknessModifier.grade,
            ThicknessModifier.surface_finish,
            ThicknessModifier.thickness
        ).all()

        for row_idx, mod in enumerate(thickness_mods, start=2):
            ws1.cell(row=row_idx, column=1, value=mod.grade)
            ws1.cell(row=row_idx, column=2, value=mod.surface_finish)
            ws1.cell(row=row_idx, column=3, value=mod.base_width)
            ws1.cell(row=row_idx, column=4, value=mod.thickness)
            ws1.cell(row=row_idx, column=5, value=mod.price_modifier)

            for col in range(1, 6):
                cell = ws1.cell(row=row_idx, column=col)
                cell.border = self.BORDER

        self._autofit_columns(ws1)

        # Arkusz 2: Modyfikatory szerokosci
        ws2 = wb.create_sheet("Modyfikatory szerokosci")

        headers = ["Gatunek", "Szerokosc (mm)", "Modyfikator PLN/kg"]
        self._write_headers(ws2, headers)

        width_mods = self.db.query(WidthModifier).order_by(
            WidthModifier.grade,
            WidthModifier.width
        ).all()

        for row_idx, mod in enumerate(width_mods, start=2):
            ws2.cell(row=row_idx, column=1, value=mod.grade or "(wszystkie)")
            ws2.cell(row=row_idx, column=2, value=mod.width)
            ws2.cell(row=row_idx, column=3, value=mod.price_modifier)

            for col in range(1, 4):
                cell = ws2.cell(row=row_idx, column=col)
                cell.border = self.BORDER

        self._autofit_columns(ws2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_all(
        self,
        categories: Optional[list[str]] = None,
        thickness_min: Optional[float] = None,
        thickness_max: Optional[float] = None,
        surface_finishes: Optional[list[str]] = None,
        only_active: bool = True,
    ) -> bytes:
        """Eksportuj wszystkie ceny do wieloarkuszowego Excela.

        Args:
            categories: Lista kategorii materialow
            thickness_min: Minimalna grubosc
            thickness_max: Maksymalna grubosc
            surface_finishes: Lista wykonczn
            only_active: Tylko aktywne ceny

        Returns:
            bytes: Zawartosc pliku Excel
        """
        wb = Workbook()

        # === Arkusz 1: Ceny bazowe ===
        ws1 = wb.active
        ws1.title = "Ceny bazowe"

        headers = [
            "Gatunek", "Nazwa materialu", "Kategoria", "Wykoncznie",
            "Grubosc (mm)", "Szerokosc (mm)", "Dlugosc (mm)", "Cena PLN/kg", "Uwagi"
        ]
        self._write_headers(ws1, headers)

        query = self.db.query(BasePrice).options(joinedload(BasePrice.material))

        if only_active:
            query = query.filter(BasePrice.is_active == True)

        if categories:
            category_enums = [MaterialCategory(c) for c in categories]
            query = query.join(Material).filter(Material.category.in_(category_enums))

        if thickness_min is not None:
            query = query.filter(BasePrice.thickness >= thickness_min)
        if thickness_max is not None:
            query = query.filter(BasePrice.thickness <= thickness_max)

        if surface_finishes:
            query = query.filter(BasePrice.surface_finish.in_(surface_finishes))

        prices = query.order_by(
            BasePrice.material_id, BasePrice.surface_finish,
            BasePrice.thickness, BasePrice.width
        ).all()

        for row_idx, price in enumerate(prices, start=2):
            material = price.material
            ws1.cell(row=row_idx, column=1, value=material.grade)
            ws1.cell(row=row_idx, column=2, value=material.name)
            ws1.cell(row=row_idx, column=3, value=material.category.value)
            ws1.cell(row=row_idx, column=4, value=price.surface_finish)
            ws1.cell(row=row_idx, column=5, value=price.thickness)
            ws1.cell(row=row_idx, column=6, value=price.width)
            ws1.cell(row=row_idx, column=7, value=price.length)
            ws1.cell(row=row_idx, column=8, value=price.price_pln_per_kg)
            ws1.cell(row=row_idx, column=9, value=price.notes or "")

            for col in range(1, 10):
                cell = ws1.cell(row=row_idx, column=col)
                cell.border = self.BORDER

        self._autofit_columns(ws1)

        # === Arkusz 2: Cennik szlifu ===
        ws2 = wb.create_sheet("Cennik szlifu")

        headers = ["Dostawca", "Granulacja", "Grubosc (mm)", "Cena PLN/kg", "Z SB", "Wariant szerokosci"]
        self._write_headers(ws2, headers)

        grinding_query = self.db.query(GrindingPrice)
        if only_active:
            grinding_query = grinding_query.filter(GrindingPrice.is_active == True)

        if thickness_min is not None:
            grinding_query = grinding_query.filter(GrindingPrice.thickness >= thickness_min)
        if thickness_max is not None:
            grinding_query = grinding_query.filter(GrindingPrice.thickness <= thickness_max)

        grinding_prices = grinding_query.order_by(
            GrindingPrice.provider, GrindingPrice.grit, GrindingPrice.thickness
        ).all()

        for row_idx, price in enumerate(grinding_prices, start=2):
            ws2.cell(row=row_idx, column=1, value=price.provider.value)
            ws2.cell(row=row_idx, column=2, value=price.grit)
            ws2.cell(row=row_idx, column=3, value=price.thickness)
            ws2.cell(row=row_idx, column=4, value=price.price_pln_per_kg)
            ws2.cell(row=row_idx, column=5, value="Tak" if price.with_sb else "Nie")
            ws2.cell(row=row_idx, column=6, value=price.width_variant or "")

            for col in range(1, 7):
                ws2.cell(row=row_idx, column=col).border = self.BORDER

        self._autofit_columns(ws2)

        # === Arkusz 3: Cennik folii ===
        ws3 = wb.create_sheet("Cennik folii")

        headers = ["Typ folii", "Grubosc (mm)", "Cena PLN/kg"]
        self._write_headers(ws3, headers)

        film_query = self.db.query(FilmPrice)
        if only_active:
            film_query = film_query.filter(FilmPrice.is_active == True)

        if thickness_min is not None:
            film_query = film_query.filter(FilmPrice.thickness >= thickness_min)
        if thickness_max is not None:
            film_query = film_query.filter(FilmPrice.thickness <= thickness_max)

        film_prices = film_query.order_by(FilmPrice.film_type, FilmPrice.thickness).all()

        for row_idx, price in enumerate(film_prices, start=2):
            ws3.cell(row=row_idx, column=1, value=price.film_type.value)
            ws3.cell(row=row_idx, column=2, value=price.thickness)
            ws3.cell(row=row_idx, column=3, value=price.price_pln_per_kg)

            for col in range(1, 4):
                ws3.cell(row=row_idx, column=col).border = self.BORDER

        self._autofit_columns(ws3)

        # === Arkusz 4: Modyfikatory grubosci ===
        ws4 = wb.create_sheet("Modyfikatory grubosci")

        headers = ["Gatunek", "Wykoncznie", "Szerokosc bazowa", "Grubosc (mm)", "Modyfikator PLN/kg"]
        self._write_headers(ws4, headers)

        thickness_mods = self.db.query(ThicknessModifier).order_by(
            ThicknessModifier.grade, ThicknessModifier.thickness
        ).all()

        for row_idx, mod in enumerate(thickness_mods, start=2):
            ws4.cell(row=row_idx, column=1, value=mod.grade)
            ws4.cell(row=row_idx, column=2, value=mod.surface_finish)
            ws4.cell(row=row_idx, column=3, value=mod.base_width)
            ws4.cell(row=row_idx, column=4, value=mod.thickness)
            ws4.cell(row=row_idx, column=5, value=mod.price_modifier)

            for col in range(1, 6):
                ws4.cell(row=row_idx, column=col).border = self.BORDER

        self._autofit_columns(ws4)

        # === Arkusz 5: Modyfikatory szerokosci ===
        ws5 = wb.create_sheet("Modyfikatory szerokosci")

        headers = ["Gatunek", "Szerokosc (mm)", "Modyfikator PLN/kg"]
        self._write_headers(ws5, headers)

        width_mods = self.db.query(WidthModifier).order_by(WidthModifier.width).all()

        for row_idx, mod in enumerate(width_mods, start=2):
            ws5.cell(row=row_idx, column=1, value=mod.grade or "(wszystkie)")
            ws5.cell(row=row_idx, column=2, value=mod.width)
            ws5.cell(row=row_idx, column=3, value=mod.price_modifier)

            for col in range(1, 4):
                ws5.cell(row=row_idx, column=col).border = self.BORDER

        self._autofit_columns(ws5)

        # Zapisz
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_base_prices_csv(
        self,
        categories: Optional[list[str]] = None,
        thickness_min: Optional[float] = None,
        thickness_max: Optional[float] = None,
        surface_finishes: Optional[list[str]] = None,
        only_active: bool = True,
    ) -> str:
        """Eksportuj ceny bazowe do CSV.

        Returns:
            str: Zawartosc pliku CSV
        """
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output, delimiter=";")

        # Naglowki
        writer.writerow([
            "Gatunek", "Nazwa materialu", "Kategoria", "Wykoncznie",
            "Grubosc (mm)", "Szerokosc (mm)", "Dlugosc (mm)", "Cena PLN/kg", "Uwagi"
        ])

        # Pobierz dane
        query = self.db.query(BasePrice).options(joinedload(BasePrice.material))

        if only_active:
            query = query.filter(BasePrice.is_active == True)

        if categories:
            category_enums = [MaterialCategory(c) for c in categories]
            query = query.join(Material).filter(Material.category.in_(category_enums))

        if thickness_min is not None:
            query = query.filter(BasePrice.thickness >= thickness_min)
        if thickness_max is not None:
            query = query.filter(BasePrice.thickness <= thickness_max)

        if surface_finishes:
            query = query.filter(BasePrice.surface_finish.in_(surface_finishes))

        prices = query.order_by(
            BasePrice.material_id, BasePrice.surface_finish,
            BasePrice.thickness, BasePrice.width
        ).all()

        for price in prices:
            material = price.material
            writer.writerow([
                material.grade,
                material.name,
                material.category.value,
                price.surface_finish,
                price.thickness,
                price.width,
                price.length,
                price.price_pln_per_kg,
                price.notes or "",
            ])

        return output.getvalue()

    def _write_headers(self, ws, headers: list[str]):
        """Zapisz sformatowane naglowki do arkusza."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = self.CENTER_ALIGN

    def _autofit_columns(self, ws):
        """Automatyczne dopasowanie szerokosci kolumn."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def get_export_filename(self, data_type: str, format: str = "xlsx") -> str:
        """Wygeneruj nazwe pliku eksportu.

        Args:
            data_type: Typ danych (base_prices, grinding, film, modifiers, all)
            format: Format pliku (xlsx, csv)

        Returns:
            str: Nazwa pliku z datą
        """
        type_names = {
            "base_prices": "ceny_bazowe",
            "grinding": "cennik_szlifu",
            "film": "cennik_folii",
            "modifiers": "modyfikatory",
            "all": "cennik_pelny",
        }
        name = type_names.get(data_type, data_type)
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        return f"{name}_{date_str}.{format}"
